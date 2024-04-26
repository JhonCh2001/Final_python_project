from openpyxl import Workbook
from openpyxl.styles import Font, Color, Alignment, PatternFill, Border
from openpyxl.drawing.image import Image
from io import BytesIO

import src.util as util

def descargar_excel(wb):
    excel_bytes = BytesIO()
    wb.save(excel_bytes)
    excel_bytes.seek(0)
    return excel_bytes

def crear_reporte_excel(df_clean):
    # Crear un nuevo libro y seleccionar la hoja activa
    wb = Workbook()
    ws = wb.active

    year = df_clean['ATENCION_AÑO'].iloc[0]

    # Añadir texto a una celda
    ws['A1'] = "Coberturas de atención de las unidades operativas de primer nivel de la dirección distrital 04D02 Montufar Bolivar Salud"

    # Establecer fuente, tamaño y color
    ws['A1'].font = Font(name='American Typewriter', size = 20, bold=True, italic=True)

    meses = df_clean['ATENCION_MES'].unique()

    ws['A4'] = f'Consultas de prevención en primer nivel de atención D04D02 Montúfar Bolívar Salud {util.mes_num(meses[0])}-{util.mes_num(meses[-1])} {year}'
    ws['A4'].font = Font(name='American Typewriter', size = 18, bold=True)

    # Añadir imagenes

    img = Image('hist.png')
    ws.add_image(img,'C6')

    etiquetas = ['Menores a 1 año','De 1 a 4 años','De 5 a 9 años','De 10 a 14 años','De 15 a 19 años','De 20 a 64 años','Mayores a 65 años']
    llaves = ['Menor_1','1_a_4','5_a_9','10_a_14','15_a_19','20_a_64','Mayor_65']

    lista = list(zip(etiquetas,llaves))

    aux = 0

    ws.column_dimensions['G'].width = 30

    for etiqueta,llave in lista:
        ws[f'G{33+aux}'] = f'{etiqueta}'
        ws[f'G{33+aux}'].font = Font(name='American Typewriter', size = 15)
        
        ws[f'H{33+aux}'] = util.edad_por_rango(0,llave,df_clean)
        ws[f'H{33+aux}'].font = Font(name='American Typewriter', size = 15)
        aux += 1

    aux_meses = 0

    for mes in meses:

        ws[f'A{44+aux_meses}'] = f'Consultas de prevención en primer nivel de atención D04D02 Montúfar Bolívar Salud {util.mes_num(mes)} {year}'
        ws[f'A{44+aux_meses}'].font = Font(name='American Typewriter', size = 18, bold=True)

        # Añadir imagenes

        img = Image(f'hist_{mes}.png')
        ws.add_image(img,f'C{46+aux_meses}')

        aux = 0

        for etiqueta,llave in lista:
            ws[f'G{73+aux+aux_meses}'] = f'{etiqueta}'
            ws[f'G{73+aux+aux_meses}'].font = Font(name='American Typewriter', size = 15)
            
            ws[f'H{73+aux+aux_meses}'] = util.edad_por_rango(mes,llave,df_clean)
            ws[f'H{73+aux+aux_meses}'].font = Font(name='American Typewriter', size = 15)
            aux += 1

        aux_meses += 40
    
    return wb